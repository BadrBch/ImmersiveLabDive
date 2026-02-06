#!/usr/bin/env python3
"""
sync_all_content.py - Sync content from DIVE25_Database.xlsx to data/content.php

This script reads the Excel database and generates a PHP array file that powers
the DIVE25 website. Run this after making changes to the Excel file.

Usage:
    python3 sync_all_content.py

Requirements:
    pip install openpyxl
"""

import os
import re
from datetime import datetime
from openpyxl import load_workbook

# Paths
SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
EXCEL_FILE = os.path.join(SCRIPT_DIR, 'DIVE25_Database.xlsx')
OUTPUT_FILE = os.path.join(SCRIPT_DIR, 'data', 'content.php')


def extract_youtube_id(url):
    """Extract YouTube video ID from URL."""
    if not url or url == '#':
        return 'dQw4w9WgXcQ'  # Default placeholder
    match = re.search(r'(?:v=|youtu\.be/)([a-zA-Z0-9_-]{11})', str(url))
    return match.group(1) if match else 'dQw4w9WgXcQ'


def escape_php_string(s):
    """Escape a string for PHP single-quoted string."""
    if s is None:
        return ''
    s = str(s)
    s = s.replace('\\', '\\\\')
    s = s.replace("'", "\\'")
    return s


def read_sheet_as_rows(ws):
    """Read worksheet and return list of non-empty rows."""
    rows = []
    for row in ws.iter_rows(values_only=True):
        if any(cell is not None for cell in row):
            rows.append(list(row))
    return rows


def parse_key_value_pairs(rows, start_idx=1):
    """Parse rows as key-value pairs until a SECTION or LIST marker."""
    result = {}
    i = start_idx
    while i < len(rows):
        row = rows[i]
        key = row[0] if row else None
        
        if key is None or str(key).startswith('SECTION:') or str(key).endswith('_START'):
            break
        
        if key and not str(key).startswith('📋'):
            value = row[1] if len(row) > 1 else None
            result[key] = value if value is not None else ''
        i += 1
    return result, i


def parse_list_until_end(rows, start_idx, end_marker, columns):
    """Parse rows as a list of dicts until end marker."""
    items = []
    i = start_idx
    while i < len(rows):
        row = rows[i]
        key = row[0] if row else None
        
        if key == end_marker:
            return items, i + 1
        
        # Skip header rows
        if key and not str(key).endswith('_START') and key not in columns:
            item = {}
            for j, col in enumerate(columns):
                item[col] = row[j] if j < len(row) and row[j] is not None else ''
            items.append(item)
        i += 1
    return items, i


def parse_simple_list_until_end(rows, start_idx, end_marker):
    """Parse rows as a simple list of strings until end marker."""
    items = []
    i = start_idx
    while i < len(rows):
        row = rows[i]
        key = row[0] if row else None
        
        if key == end_marker:
            return items, i + 1
        
        if key and not str(key).endswith('_START'):
            items.append(str(key))
        i += 1
    return items, i


def parse_home_sheet(ws):
    """Parse HOME sheet."""
    rows = read_sheet_as_rows(ws)
    data = {}
    
    for i, row in enumerate(rows):
        key = row[0] if row else None
        value = row[1] if len(row) > 1 else None
        
        if key and not str(key).startswith('📋') and not str(key).startswith('←'):
            if key == 'PILLARS_START':
                # Parse 3 pillars
                pillars = []
                for j in range(1, 4):
                    if i + j < len(rows):
                        r = rows[i + j]
                        pillars.append({
                            'title': r[0] if r[0] else '',
                            'description': r[1] if len(r) > 1 and r[1] else ''
                        })
                data['association_pillars'] = pillars
            elif key == 'RESEARCH_PILLARS_START':
                # Parse research pillars (multi-column)
                pillars = []
                j = i + 2  # Skip header
                while j < len(rows) and rows[j][0] != 'RESEARCH_PILLARS_END':
                    r = rows[j]
                    if r[0]:
                        pillar = {'name': r[0], 'items': []}
                        # Items are in additional columns
                        for k in range(1, len(r)):
                            if r[k]:
                                pillar['items'].append(str(r[k]))
                        pillars.append(pillar)
                    j += 1
                data['research_pillars'] = pillars
            elif value is not None:
                data[key] = value
    
    return data


def parse_about_sheet(ws):
    """Parse ABOUT sheet."""
    rows = read_sheet_as_rows(ws)
    data = {}
    
    for row in rows:
        key = row[0] if row else None
        value = row[1] if len(row) > 1 else None
        
        if key and not str(key).startswith('📋') and not str(key).startswith('←'):
            if value is not None:
                data[key] = value
    
    return data


def parse_education_sheet(ws):
    """Parse EDUCATION sheet."""
    rows = read_sheet_as_rows(ws)
    data = {
        'pillars': [],
        'first_year_actions': [],
        'contacts': [],
        'inventory': [],
        'upgrades': [],
    }
    
    i = 0
    while i < len(rows):
        row = rows[i]
        key = row[0] if row else None
        value = row[1] if len(row) > 1 else None
        
        if key is None:
            i += 1
            continue
            
        if key.startswith('education_') or key.startswith('hardware_') or key.startswith('first_year_') or key.startswith('upgrades_') or key.startswith('resources_'):
            data[key] = value
        elif key.startswith('pillar_') and key.endswith('_title'):
            num = key.split('_')[1]
            desc_key = f'pillar_{num}_desc'
            for r in rows:
                if r[0] == desc_key:
                    data['pillars'].append({
                        'title': value,
                        'description': r[1] if len(r) > 1 else ''
                    })
                    break
        elif key.startswith('action_'):
            data['first_year_actions'].append(value)
        elif key.startswith('contact_') and key.endswith('_label'):
            num = key.split('_')[1]
            value_key = f'contact_{num}_value'
            for r in rows:
                if r[0] == value_key:
                    data['contacts'].append({
                        'label': value,
                        'value': r[1] if len(r) > 1 else ''
                    })
                    break
        elif key == 'HARDWARE_LIST_START':
            i += 1
            while i < len(rows) and rows[i][0] != 'HARDWARE_LIST_END':
                r = rows[i]
                if r[0]:
                    data['inventory'].append({
                        'equipment': r[0],
                        'quantity': str(r[1]) if len(r) > 1 and r[1] else '',
                        'entity': '',
                        'contact': ''
                    })
                i += 1
        elif key == 'UPGRADE_LIST_START':
            i += 1
            while i < len(rows) and rows[i][0] != 'UPGRADE_LIST_END':
                r = rows[i]
                if r[0]:
                    data['upgrades'].append({
                        'item': r[0],
                        'priority': r[1] if len(r) > 1 and r[1] else 'Medium',
                        'quantity': 1,
                        'unit_cost': 0.00,
                        'total_cost': 0.00,
                        'link': '#'
                    })
                i += 1
        i += 1
    
    return data


def parse_research_sheet(ws):
    """Parse RESEARCH sheet."""
    rows = read_sheet_as_rows(ws)
    data = {
        'pillars': [],
        'resources': [],
        'partners': []
    }
    
    i = 0
    while i < len(rows):
        row = rows[i]
        key = row[0] if row else None
        value = row[1] if len(row) > 1 else None
        
        if key is None:
            i += 1
            continue
        
        if key.startswith('research_') or key.startswith('partners_') or key.startswith('resources_'):
            data[key] = value
        elif key == 'RESEARCH_PILLARS_START':
            i += 2  # Skip header
            while i < len(rows) and rows[i][0] != 'RESEARCH_PILLARS_END':
                r = rows[i]
                if r[0]:
                    items = [str(r[j]) for j in range(1, len(r)) if r[j]]
                    data['pillars'].append({'name': r[0], 'items': items})
                i += 1
        elif key == 'RESEARCH_RESOURCES_START':
            i += 2  # Skip header
            while i < len(rows) and rows[i][0] != 'RESEARCH_RESOURCES_END':
                r = rows[i]
                if r[0]:
                    data['resources'].append({
                        'label': r[0],
                        'description': r[1] if len(r) > 1 and r[1] else '',
                        'path': r[2] if len(r) > 2 and r[2] else ''
                    })
                i += 1
        i += 1
    
    return data


def parse_projects_sheet(ws):
    """Parse PROJECTS sheet."""
    rows = read_sheet_as_rows(ws)
    data = {
        'videos': [],
        'categories': []
    }
    
    i = 0
    while i < len(rows):
        row = rows[i]
        key = row[0] if row else None
        value = row[1] if len(row) > 1 else None
        
        if key is None:
            i += 1
            continue
        
        if key.startswith('projects_') or key.startswith('showcase_') or key.startswith('submit_') or key.startswith('channel_') or key.startswith('categories_') or key.startswith('tab'):
            data[key] = value
        elif key == 'PROJECTS_LIST_START':
            i += 2  # Skip headers
            while i < len(rows):
                r = rows[i]
                if r[0] == 'PROJECTS_LIST_END' or (r[0] and r[0].startswith('SECTION:')):
                    break
                if r[0] and not r[0].startswith('Title'):
                    youtube_url = r[3] if len(r) > 3 and r[3] else '#'
                    data['videos'].append({
                        'title': r[0],
                        'category': r[1] if len(r) > 1 and r[1] else '',
                        'description': r[2] if len(r) > 2 and r[2] else '',
                        'youtube_url': youtube_url,
                        'youtube_id': extract_youtube_id(youtube_url)
                    })
                i += 1
            continue
        elif key == 'CATEGORIES_LIST_START':
            i += 1  # Skip header
            while i < len(rows):
                r = rows[i]
                if r[0] == 'CATEGORIES_LIST_END' or r[0] is None:
                    break
                if r[0] and not r[0].startswith('Title'):
                    data['categories'].append({
                        'title': r[0],
                        'description': r[1] if len(r) > 1 and r[1] else '',
                        'icon': r[2] if len(r) > 2 and r[2] else 'puzzle'
                    })
                i += 1
            continue
        i += 1
    
    return data


def parse_events_sheet(ws):
    """Parse EVENTS sheet."""
    rows = read_sheet_as_rows(ws)
    data = {
        'highlights': [],
        'why_attend': [],
        'timeline': [],
        'participation_steps': [],
        'contacts': [],
        'calendar_events': []
    }
    
    i = 0
    while i < len(rows):
        row = rows[i]
        key = row[0] if row else None
        value = row[1] if len(row) > 1 else None
        
        if key is None:
            i += 1
            continue
        
        if key.startswith('events_') or key.startswith('calendar_') or key.startswith('organizer_') or key.startswith('cta_'):
            data[key] = value
        elif key == 'CALENDAR_EVENTS_START':
            i += 1
            while i < len(rows) and rows[i][0] != 'CALENDAR_EVENTS_END':
                r = rows[i]
                if r[0]:
                    data['calendar_events'].append({
                        'date': r[0],
                        'year': r[1] if len(r) > 1 else '',
                        'title': r[2] if len(r) > 2 else '',
                        'location': r[3] if len(r) > 3 else '',
                        'description': r[4] if len(r) > 4 else ''
                    })
                i += 1
        elif key == 'HIGHLIGHTS_START':
            i += 1
            while i < len(rows) and rows[i][0] != 'HIGHLIGHTS_END':
                if rows[i][0]:
                    data['highlights'].append(rows[i][0])
                i += 1
        elif key == 'WHY_ATTEND_START':
            i += 1
            while i < len(rows) and rows[i][0] != 'WHY_ATTEND_END':
                if rows[i][0]:
                    data['why_attend'].append(rows[i][0])
                i += 1
        elif key == 'PARTICIPATION_START':
            i += 1
            while i < len(rows) and rows[i][0] != 'PARTICIPATION_END':
                r = rows[i]
                if r[0]:
                    data['participation_steps'].append({
                        'text': r[0],
                        'link': r[1] if len(r) > 1 and r[1] else '#',
                        'link_label': r[2] if len(r) > 2 and r[2] else ''
                    })
                i += 1
        elif key == 'CONTACTS_START':
            i += 1
            while i < len(rows) and rows[i][0] != 'CONTACTS_END':
                r = rows[i]
                if r[0]:
                    data['contacts'].append({
                        'label': r[0],
                        'value': r[1] if len(r) > 1 and r[1] else ''
                    })
                i += 1
        i += 1
    
    return data


def parse_contacts_sheet(ws):
    """Parse CONTACTS sheet."""
    rows = read_sheet_as_rows(ws)
    data = {
        'team': [],
        'social': []
    }
    
    i = 0
    while i < len(rows):
        row = rows[i]
        key = row[0] if row else None
        value = row[1] if len(row) > 1 else None
        
        if key is None:
            i += 1
            continue
        
        if key.startswith('contact_') or key.startswith('social_'):
            if key.startswith('social_') and value:
                platform = key.replace('social_', '')
                data['social'].append({'platform': platform, 'url': value})
            else:
                data[key] = value
        elif key == 'TEAM_LIST_START':
            i += 2  # Skip headers
            while i < len(rows) and rows[i][0] != 'TEAM_LIST_END':
                r = rows[i]
                if r[0] and r[0] != 'First Name':
                    data['team'].append({
                        'firstname': r[0],
                        'lastname': r[1] if len(r) > 1 and r[1] else '',
                        'job': r[2] if len(r) > 2 and r[2] else '',
                        'email': r[3] if len(r) > 3 and r[3] else '',
                        'image': r[4] if len(r) > 4 and r[4] else ''
                    })
                i += 1
        i += 1
    
    return data


def format_php_value(value, indent=2):
    """Format a value for PHP output."""
    indent_str = '    ' * indent
    
    if value is None:
        return "''"
    elif isinstance(value, bool):
        return 'true' if value else 'false'
    elif isinstance(value, (int, float)):
        return str(value)
    elif isinstance(value, list):
        if not value:
            return '[]'
        items = []
        for item in value:
            items.append(format_php_value(item, indent + 1))
        return '[\n' + ',\n'.join(f'{indent_str}    {item}' for item in items) + f'\n{indent_str}]'
    elif isinstance(value, dict):
        items = []
        for k, v in value.items():
            items.append(f"'{escape_php_string(k)}' => {format_php_value(v, indent + 1)}")
        return '[' + ', '.join(items) + ']'
    else:
        return f"'{escape_php_string(str(value))}'"


def generate_php_content(data):
    """Generate the full PHP content file."""
    lines = [
        '<?php',
        '/**',
        ' * Central content data file - AUTO-GENERATED from Excel',
        ' * DO NOT EDIT MANUALLY - Use DIVE25_Database.xlsx and run sync_all_content.py',
        ' */',
        '',
        'return ['
    ]
    
    # HOME section
    home = data.get('home', {})
    lines.append("    'home' => [")
    lines.append(f"        'hero_title' => '{escape_php_string(home.get('hero_title', 'What is Immersive Learning?'))}',")
    lines.append(f"        'registration_title' => '{escape_php_string(home.get('registration_title', 'Immersive LAB WORKSHOP'))}',")
    lines.append(f"        'registration_text' => '{escape_php_string(home.get('registration_text', ''))}',")
    lines.append(f"        'registration_btn' => '{escape_php_string(home.get('registration_btn', 'ACCESS REGISTRATION DETAILS'))}',")
    lines.append("        ")
    lines.append(f"        'association_title' => '{escape_php_string(home.get('association_title', 'Virtual Vinci Student Association'))}',")
    lines.append(f"        'association_tagline' => '{escape_php_string(home.get('association_tagline', ''))}',")
    lines.append(f"        'association_summary' => '{escape_php_string(home.get('association_summary', ''))}',")
    
    # Association pillars
    pillars = home.get('association_pillars', [])
    lines.append("        'association_pillars' => [")
    for p in pillars:
        lines.append(f"            ['title' => '{escape_php_string(p.get('title', ''))}', 'description' => '{escape_php_string(p.get('description', ''))}'],")
    lines.append("        ],")
    lines.append("        ")
    
    lines.append(f"        'research_title' => '{escape_php_string(home.get('research_title', 'Vision & Immersion Research Group'))}',")
    lines.append(f"        'research_intro' => '{escape_php_string(home.get('research_intro', ''))}',")
    lines.append(f"        'research_contact' => '{escape_php_string(home.get('research_contact', ''))}',")
    
    # Research pillars
    research_pillars = home.get('research_pillars', [])
    lines.append("        'research_pillars' => [")
    for p in research_pillars:
        items_str = ', '.join(f"'{escape_php_string(item)}'" for item in p.get('items', []))
        lines.append(f"            ['name' => '{escape_php_string(p.get('name', ''))}', 'items' => [{items_str}, ]],")
    lines.append("        ]")
    lines.append("    ],")
    lines.append("")
    
    # ASSOCIATION/EDUCATION section
    edu = data.get('education', {})
    lines.append("    'association' => [")
    lines.append(f"        'page_title' => '{escape_php_string(edu.get('education_page_title', 'EDUCATION & TRAINING'))}',")
    lines.append(f"        'title' => '{escape_php_string(edu.get('education_main_title', 'VIRTUAL VINCI STUDENT ASSOCIATION'))}',")
    lines.append(f"        'tagline' => '',")
    lines.append(f"        'summary' => '{escape_php_string(edu.get('education_description', ''))}',")
    lines.append(f"        'first_year_title' => '{escape_php_string(edu.get('first_year_title', 'LEARNING PROGRAMS & ACTIVITIES'))}',")
    lines.append(f"        'resources_title' => '{escape_php_string(edu.get('resources_title', 'EDUCATIONAL RESOURCES'))}',")
    lines.append(f"        'resources_intro' => '{escape_php_string(edu.get('resources_intro', ''))}',")
    
    # Pillars
    lines.append("        'pillars' => [")
    for p in edu.get('pillars', []):
        lines.append(f"            ['title' => '{escape_php_string(p.get('title', ''))}', 'description' => '{escape_php_string(p.get('description', ''))}'],")
    lines.append("        ],")
    
    # First year actions
    lines.append("        'first_year_actions' => [")
    for action in edu.get('first_year_actions', []):
        lines.append(f"            '{escape_php_string(action)}',")
    lines.append("        ],")
    
    # Contacts
    lines.append("        'contacts' => [")
    for c in edu.get('contacts', []):
        lines.append(f"            ['label' => '{escape_php_string(c.get('label', ''))}', 'value' => '{escape_php_string(c.get('value', ''))}'],")
    lines.append("        ],")
    lines.append("        'resources' => []")
    lines.append("    ],")
    lines.append("")
    
    # RESEARCH section
    research = data.get('research', {})
    lines.append("    'research' => [")
    lines.append(f"        'page_title' => '{escape_php_string(research.get('research_page_title', 'RESEARCH'))}',")
    lines.append(f"        'title' => '{escape_php_string(research.get('research_main_title', 'VISION & IMMERSION RESEARCH GROUP'))}',")
    lines.append(f"        'intro' => '{escape_php_string(research.get('research_intro', ''))}',")
    lines.append(f"        'contact' => '{escape_php_string(research.get('research_contact', ''))}',")
    lines.append(f"        'pillars_title' => '{escape_php_string(research.get('pillars_title', 'ACTIVE RESEARCH PROJECTS'))}',")
    lines.append(f"        'partners_title' => '{escape_php_string(research.get('partners_title', 'XR ECOSYSTEM & PARTNERS'))}',")
    lines.append(f"        'partners_intro' => '{escape_php_string(research.get('partners_intro', ''))}',")
    lines.append(f"        'resources_title' => '{escape_php_string(research.get('resources_title', 'RESEARCH RESOURCES'))}',")
    lines.append(f"        'resources_intro' => '{escape_php_string(research.get('resources_intro', ''))}',")
    
    lines.append("        'pillars' => [")
    for p in research.get('pillars', []):
        items_str = ', '.join(f"'{escape_php_string(item)}'" for item in p.get('items', []))
        lines.append(f"            ['name' => '{escape_php_string(p.get('name', ''))}', 'items' => [{items_str}, ]],")
    lines.append("        ],")
    
    lines.append("        'resources' => [")
    for r in research.get('resources', []):
        lines.append(f"            ['label' => '{escape_php_string(r.get('label', ''))}', 'description' => '{escape_php_string(r.get('description', ''))}', 'path' => '{escape_php_string(r.get('path', ''))}'],")
    lines.append("        ]")
    lines.append("    ],")
    lines.append("")
    
    # EVENTS section
    events = data.get('events', {})
    lines.append("    'events' => [")
    lines.append("        'featured' => [")
    lines.append(f"            'title' => '{escape_php_string(events.get('events_main_title', 'PIDS'))}',")
    lines.append(f"            'tagline' => '{escape_php_string(events.get('events_tagline', ''))}',")
    lines.append(f"            'date_label' => '{escape_php_string(events.get('events_date_label', ''))}',")
    lines.append(f"            'location' => '{escape_php_string(events.get('events_location', ''))}',")
    lines.append(f"            'summary' => '{escape_php_string(events.get('events_summary', ''))}',")
    
    lines.append("            'highlights' => [")
    for h in events.get('highlights', []):
        lines.append(f"                '{escape_php_string(h)}',")
    lines.append("            ],")
    
    lines.append("            'why_attend' => [")
    for w in events.get('why_attend', []):
        lines.append(f"                '{escape_php_string(w)}',")
    lines.append("            ],")
    
    lines.append("            'timeline' => [")
    lines.append("            ],")
    
    lines.append("            'participation_steps' => [")
    for p in events.get('participation_steps', []):
        lines.append(f"                ['text' => '{escape_php_string(p.get('text', ''))}', 'link' => '{escape_php_string(p.get('link', '#'))}', 'link_label' => '{escape_php_string(p.get('link_label', ''))}'],")
    lines.append("            ],")
    
    lines.append("            'contacts' => [")
    for c in events.get('contacts', []):
        lines.append(f"                ['label' => '{escape_php_string(c.get('label', ''))}', 'value' => '{escape_php_string(c.get('value', ''))}'],")
    lines.append("            ],")
    
    lines.append(f"            'cta_title' => '{escape_php_string(events.get('cta_title', 'FOLLOW OUR IMMERSIVE LAB JOURNEY'))}',")
    lines.append(f"            'cta_text' => '{escape_php_string(events.get('cta_text', ''))}',")
    lines.append(f"            'cta_btn_text' => '{escape_php_string(events.get('cta_btn_text', 'VISIT OUR YOUTUBE CHANNEL'))}',")
    lines.append(f"            'cta_link' => '{escape_php_string(events.get('cta_link', '#'))}'")
    lines.append("        ]")
    lines.append("    ],")
    lines.append("")
    
    # HARDWARE section
    lines.append("    'hardware' => [")
    lines.append(f"        'title' => '{escape_php_string(edu.get('hardware_title', 'XR HARDWARE LAB'))}',")
    lines.append(f"        'intro' => '{escape_php_string(edu.get('hardware_intro', ''))}',")
    lines.append(f"        'upgrades_title' => '{escape_php_string(edu.get('upgrades_title', 'UPCOMING EQUIPMENT INVESTMENTS'))}',")
    
    lines.append("        'inventory' => [")
    for item in edu.get('inventory', []):
        lines.append(f"            ['equipment' => '{escape_php_string(item.get('equipment', ''))}', 'quantity' => '{escape_php_string(item.get('quantity', ''))}', 'entity' => '', 'contact' => ''],")
    lines.append("        ],")
    
    lines.append("        'upgrades' => [")
    for u in edu.get('upgrades', []):
        lines.append(f"            ['item' => '{escape_php_string(u.get('item', ''))}', 'priority' => '{escape_php_string(u.get('priority', 'Medium'))}', 'quantity' => 1, 'unit_cost' => 0.00, 'total_cost' => 0.00, 'link' => '#'],")
    lines.append("        ],")
    lines.append("        'total_investment' => 0.00")
    lines.append("    ],")
    lines.append("")
    
    # VENDORS section (empty)
    lines.append("    'vendors' => [")
    lines.append("    ],")
    lines.append("")
    
    # DOWNLOADS section
    lines.append("    'downloads' => [")
    # Get from downloads in research resources + some defaults
    for r in research.get('resources', []):
        lines.append(f"        ['label' => '{escape_php_string(r.get('label', ''))}', 'description' => '{escape_php_string(r.get('description', ''))}', 'path' => '{escape_php_string(r.get('path', ''))}'],")
    lines.append("    ],")
    lines.append("")
    
    # PROJECTS section
    projects = data.get('projects', {})
    lines.append("    'projects' => [")
    lines.append(f"        'page_title' => '{escape_php_string(projects.get('projects_page_title', 'PROJECTS'))}',")
    lines.append(f"        'title' => '{escape_php_string(projects.get('projects_main_title', 'XR PROJECTS SHOWCASE'))}',")
    lines.append(f"        'intro' => '{escape_php_string(projects.get('projects_intro', ''))}',")
    lines.append(f"        'showcase_title' => '{escape_php_string(projects.get('showcase_title', 'SHOWCASE YOUR XR PROJECT'))}',")
    lines.append(f"        'showcase_text' => '{escape_php_string(projects.get('showcase_text', 'Have you developed...'))}',")
    lines.append(f"        'submit_btn_text' => '{escape_php_string(projects.get('submit_btn_text', 'SUBMIT YOUR PROJECT'))}',")
    lines.append(f"        'submit_link' => '{escape_php_string(projects.get('submit_link', '#'))}',")
    lines.append(f"        'channel_btn_text' => '{escape_php_string(projects.get('channel_btn_text', 'VISIT OUR YOUTUBE CHANNEL'))}',")
    lines.append(f"        'channel_link' => '{escape_php_string(projects.get('channel_link', 'https://www.youtube.com/@DeVinciImmersiveLab'))}',")
    lines.append(f"        'categories_title' => '{escape_php_string(projects.get('categories_title', 'PROJECT CATEGORIES'))}',")
    lines.append(f"        'categories_intro' => '{escape_php_string(projects.get('categories_intro', ''))}',")
    
    lines.append("        'videos' => [")
    for v in projects.get('videos', []):
        lines.append(f"            ['title' => '{escape_php_string(v.get('title', ''))}', 'category' => '{escape_php_string(v.get('category', ''))}', 'description' => '{escape_php_string(v.get('description', ''))}', 'youtube_url' => '{escape_php_string(v.get('youtube_url', '#'))}', 'youtube_id' => '{escape_php_string(v.get('youtube_id', ''))}'],")
    lines.append("        ],")
    
    lines.append(f"        'tab1_title' => '{escape_php_string(projects.get('tab1_title', 'Immersive Gaming '))}',")
    lines.append(f"        'tab2_title' => '{escape_php_string(projects.get('tab2_title', 'Professional Simulation'))}',")
    lines.append(f"        'tab3_title' => '{escape_php_string(projects.get('tab3_title', 'Healthcare & Medical Training'))}',")
    
    lines.append("        'categories' => [")
    for c in projects.get('categories', []):
        lines.append(f"            ['title' => '{escape_php_string(c.get('title', ''))}', 'description' => '{escape_php_string(c.get('description', ''))}', 'icon' => '{escape_php_string(c.get('icon', 'puzzle'))}'],")
    lines.append("        ]")
    lines.append("    ],")
    lines.append("")
    
    # CONTACT section
    contacts = data.get('contacts', {})
    lines.append("    'contact' => [")
    lines.append(f"        'page_title' => '{escape_php_string(contacts.get('contact_page_title', 'CONTACT'))}',")
    lines.append(f"        'address' => '{escape_php_string(contacts.get('contact_address_long', ''))}',")
    lines.append(f"        'email_intro' => '{escape_php_string(contacts.get('contact_email_intro', ''))}',")
    lines.append(f"        'email' => '{escape_php_string(contacts.get('contact_email', ''))}',")
    
    lines.append("        'social' => [")
    for s in contacts.get('social', []):
        lines.append(f"            ['platform' => '{escape_php_string(s.get('platform', ''))}', 'url' => '{escape_php_string(s.get('url', ''))}'],")
    lines.append("        ],")
    
    lines.append("        'team' => [")
    for t in contacts.get('team', []):
        lines.append(f"            ['firstname' => '{escape_php_string(t.get('firstname', ''))}', 'lastname' => '{escape_php_string(t.get('lastname', ''))}', 'job' => '{escape_php_string(t.get('job', ''))}', 'email' => '{escape_php_string(t.get('email', ''))}', 'image' => '{escape_php_string(t.get('image', ''))}'],")
    lines.append("        ]")
    lines.append("    ]")
    
    lines.append("];")
    lines.append("")
    
    return '\n'.join(lines)


def main():
    print(f"📊 Loading {EXCEL_FILE}...")
    wb = load_workbook(EXCEL_FILE, data_only=True)
    
    print("📝 Parsing sheets...")
    data = {}
    
    # Parse each sheet
    if 'HOME' in wb.sheetnames:
        data['home'] = parse_home_sheet(wb['HOME'])
        print("   ✓ HOME")
    
    if 'ABOUT' in wb.sheetnames:
        data['about'] = parse_about_sheet(wb['ABOUT'])
        print("   ✓ ABOUT")
    
    if 'EDUCATION' in wb.sheetnames:
        data['education'] = parse_education_sheet(wb['EDUCATION'])
        print("   ✓ EDUCATION")
    
    if 'RESEARCH' in wb.sheetnames:
        data['research'] = parse_research_sheet(wb['RESEARCH'])
        print("   ✓ RESEARCH")
    
    if 'PROJECTS' in wb.sheetnames:
        data['projects'] = parse_projects_sheet(wb['PROJECTS'])
        print("   ✓ PROJECTS")
    
    if 'EVENTS' in wb.sheetnames:
        data['events'] = parse_events_sheet(wb['EVENTS'])
        print("   ✓ EVENTS")
    
    if 'CONTACTS' in wb.sheetnames:
        data['contacts'] = parse_contacts_sheet(wb['CONTACTS'])
        print("   ✓ CONTACTS")
    
    print(f"📄 Generating {OUTPUT_FILE}...")
    php_content = generate_php_content(data)
    
    # Ensure output directory exists
    os.makedirs(os.path.dirname(OUTPUT_FILE), exist_ok=True)
    
    with open(OUTPUT_FILE, 'w', encoding='utf-8') as f:
        f.write(php_content)
    
    print(f"✅ Successfully synced content to {OUTPUT_FILE}")
    print(f"   Generated {len(php_content)} bytes")


if __name__ == '__main__':
    main()
